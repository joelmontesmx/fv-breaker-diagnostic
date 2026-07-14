"""Extraction engine for the Front View BOM Extractor web app.

This version keeps the validated breaker extraction behavior and adds Box BOM
and Interior BOM extraction. All catalog numbers are checked against the same
part-number mapping file used by the breaker workflow.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Callable, Iterable, Mapping, Sequence

import pandas as pd
from pypdf import PdfReader


BREAKER_SECTION = "Breaker BOM"
INTERIOR_SECTION = "Interior BOM"
BOX_SECTION = "Box BOM"
ALL_SECTIONS = [BREAKER_SECTION, INTERIOR_SECTION, BOX_SECTION]

BREAKER_DETAIL_FIELDS = [
    "X Space",
    "Trip Amps",
    "Sensor Amps",
    "Trip Unit",
    "Poles",
    "Phases Used",
    "Lug Cable Size / Neut Sensor / Grnd Fault Cable",
    "Notes",
]


@dataclass(frozen=True)
class CatalogMatch:
    catalog_number: str
    alternate_sku: str
    abb_part_number: str
    detected_code: str
    mapping_found: bool


@dataclass(frozen=True)
class UploadedPDF:
    name: str
    data: bytes


ProgressCallback = Callable[[int, int, str], None]


def normalize_code(value: object) -> str:
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip().upper()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def clean_token(value: str) -> str:
    return value.strip().strip("-.,;:()[]{}<>\"'").upper()


def parse_so_item(filename: str) -> tuple[str, str, str]:
    stem = Path(filename).stem
    stem = re.sub(r"\s*\(\d+\)$", "", stem).strip()

    match = re.search(r"(?<!\d)(\d{8,10})[-_](\d{1,4})(?!\d)", stem)
    if not match:
        return "NOT FOUND", "NOT FOUND", stem

    sales_order = match.group(1)
    item = match.group(2)
    return sales_order, item, f"{sales_order}-{item}"


def read_mapping_dataframe(data: bytes) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(data))
    if df.shape[1] < 2:
        raise ValueError(
            "The part-number mapping file must contain at least two columns: "
            "ABB Part Number and Alternate SKU."
        )

    df = df.iloc[:, :2].copy()
    df.columns = ["ABB Part Number", "Alternate SKU"]
    df["ABB Part Number"] = df["ABB Part Number"].map(normalize_code)
    df["Alternate SKU"] = df["Alternate SKU"].map(normalize_code)
    df = df[(df["ABB Part Number"] != "") & (df["Alternate SKU"] != "")]
    return df.drop_duplicates(
        subset=["ABB Part Number", "Alternate SKU"], keep="last"
    ).reset_index(drop=True)


def load_mappings_from_bytes(
    data: bytes,
    manual_mappings: Mapping[str, str] | None = None,
) -> tuple[dict[str, str], dict[str, str]]:
    df = read_mapping_dataframe(data)

    if manual_mappings:
        manual_rows = [
            {
                "ABB Part Number": normalize_code(abb_part_number),
                "Alternate SKU": normalize_code(alternate_sku),
            }
            for alternate_sku, abb_part_number in manual_mappings.items()
            if normalize_code(alternate_sku) and normalize_code(abb_part_number)
        ]
        if manual_rows:
            df = pd.concat([df, pd.DataFrame(manual_rows)], ignore_index=True)
            df = df.drop_duplicates(
                subset=["ABB Part Number", "Alternate SKU"], keep="last"
            )

    alt_to_abb = dict(zip(df["Alternate SKU"], df["ABB Part Number"]))
    abb_to_alt = dict(zip(df["ABB Part Number"], df["Alternate SKU"]))
    return alt_to_abb, abb_to_alt


def map_catalog_number(
    value: object,
    alt_to_abb: dict[str, str],
    abb_to_alt: dict[str, str],
) -> CatalogMatch:
    detected = normalize_code(value)
    if not detected:
        return CatalogMatch("", "", "", "", False)

    if detected in alt_to_abb:
        return CatalogMatch(
            catalog_number=detected,
            alternate_sku=detected,
            abb_part_number=alt_to_abb[detected],
            detected_code=detected,
            mapping_found=True,
        )

    if detected in abb_to_alt:
        return CatalogMatch(
            catalog_number=detected,
            alternate_sku=abb_to_alt[detected],
            abb_part_number=detected,
            detected_code=detected,
            mapping_found=True,
        )

    # Some FV tables can already contain ABB-style part numbers. Keep the code
    # as the ABB Part Number only when the prefix strongly indicates that case.
    if detected.startswith(("1SDX", "1SQA")):
        return CatalogMatch(
            catalog_number=detected,
            alternate_sku="",
            abb_part_number=detected,
            detected_code=detected,
            mapping_found=True,
        )

    return CatalogMatch(
        catalog_number=detected,
        alternate_sku=detected,
        abb_part_number="",
        detected_code=detected,
        mapping_found=False,
    )


def is_panel_marks_page(text: str) -> bool:
    return "PANEL MARKS" in text.upper()


def is_bom_page(text: str) -> bool:
    upper = text.upper()
    return "BOX BOM" in upper or "INTERIOR BOM" in upper


def is_breaker_table_page(text: str) -> bool:
    upper = text.upper()
    strong_headers = [
        "BREAKER CATALOG NUMBER",
        "SPACE BREAKER CATALOG NUMBER",
        "NEXT DEVICE INFORMATION",
    ]
    if any(header in upper for header in strong_headers):
        return True

    required_fragments = ["TRIP UNIT", "POLES", "LUG CABLE", "NOTES"]
    return "CATALOG NUMBER" in upper and sum(
        fragment in upper for fragment in required_fragments
    ) >= 3


def panel_label(index: int, total: int) -> str:
    return "Single Panel" if total <= 1 else f"Multi-Section Panel {index}"


def group_breaker_pages(page_texts: list[str]) -> list[list[int]]:
    panel_pages = [
        index for index, text in enumerate(page_texts) if is_panel_marks_page(text)
    ]
    breaker_pages = [
        index for index, text in enumerate(page_texts) if is_breaker_table_page(text)
    ]

    if panel_pages:
        groups: list[list[int]] = []
        for position, panel_index in enumerate(panel_pages):
            next_panel = (
                panel_pages[position + 1]
                if position + 1 < len(panel_pages)
                else len(page_texts)
            )
            group = [
                page_index
                for page_index in breaker_pages
                if panel_index < page_index < next_panel
            ]
            if not group and panel_index + 1 < len(page_texts):
                group = [panel_index + 1]
            if group:
                groups.append(group)
        if groups:
            return groups

    if breaker_pages:
        return [breaker_pages]

    page_count = len(page_texts)
    if page_count == 1:
        return [[0]]
    if page_count == 2:
        return [[1]]
    if page_count == 3:
        return [[1, 2]]
    return [[1]]


def group_bom_pages(page_texts: list[str]) -> list[int]:
    return [index for index, text in enumerate(page_texts) if is_bom_page(text)]


def identify_spacer_notes(text: str) -> set[str]:
    notes: set[str] = set()
    for line in text.splitlines():
        lower = line.lower()
        if "breaker space" in lower and "breaker is not included" in lower:
            match = re.match(r"\s*(\d+)\s*[.)-]?", line)
            if match:
                notes.add(match.group(1))
    return notes


def note_at_end_of_row(line: str) -> str:
    parts = line.strip().split()
    if not parts:
        return ""
    last = clean_token(parts[-1])
    return last if last.isdigit() else ""


def looks_like_full_breaker_code(token: str) -> bool:
    if not token or "SPACE" in token:
        return False

    patterns = [
        r"^XT[1-7][A-Z0-9]{8,}$",
        r"^TEY[A-Z0-9]{7,}$",
        r"^NEF[A-Z0-9]{7,}$",
        r"^1SDX[A-Z0-9]{5,}$",
        r"^1SQA[A-Z0-9]{5,}$",
    ]
    return any(re.fullmatch(pattern, token) for pattern in patterns)


def find_breaker_matches_in_line(
    line: str,
    alt_to_abb: dict[str, str],
    abb_to_alt: dict[str, str],
) -> list[CatalogMatch]:
    matches: list[CatalogMatch] = []

    for raw_token in line.split():
        token = clean_token(raw_token)
        if not token:
            continue

        if token in alt_to_abb or token in abb_to_alt or looks_like_full_breaker_code(token):
            matches.append(map_catalog_number(token, alt_to_abb, abb_to_alt))

    # A table row should only contain one breaker catalog number, but keeping a
    # unique list protects against duplicated tokens caused by PDF text extraction.
    unique: list[CatalogMatch] = []
    seen: set[str] = set()
    for match in matches:
        if match.detected_code and match.detected_code not in seen:
            unique.append(match)
            seen.add(match.detected_code)
    return unique


def _is_phase_token(token: str) -> bool:
    normalized = token.replace(",", "").upper()
    return bool(re.fullmatch(r"[ABC]+", normalized)) and 1 <= len(normalized) <= 3


def parse_breaker_row_details(line: str, breaker_code: str) -> dict[str, str]:
    details = {field: "" for field in BREAKER_DETAIL_FIELDS}
    parts = line.strip().split()
    code = normalize_code(breaker_code)
    code_position = None
    for index, part in enumerate(parts):
        if clean_token(part) == code:
            code_position = index
            break

    if code_position is None:
        return details

    if code_position > 0:
        details["X Space"] = clean_token(parts[code_position - 1])

    after_code = parts[code_position + 1 :]
    if len(after_code) >= 1:
        details["Trip Amps"] = clean_token(after_code[0])
    if len(after_code) >= 2:
        details["Sensor Amps"] = clean_token(after_code[1])

    pole_index = None
    for index in range(2, len(after_code) - 1):
        if re.fullmatch(r"[1-4]", clean_token(after_code[index])) and _is_phase_token(
            clean_token(after_code[index + 1])
        ):
            pole_index = index
            break

    if pole_index is None:
        return details

    details["Trip Unit"] = " ".join(after_code[2:pole_index]).strip()
    details["Poles"] = clean_token(after_code[pole_index])
    details["Phases Used"] = clean_token(after_code[pole_index + 1])

    remaining = after_code[pole_index + 2 :]
    if remaining:
        # Notes usually appear as the final numeric token after the lug-cable
        # field. The lug-cable field itself normally ends in AWG, MCM, or a lug
        # description, so this rule preserves the validated spacer-note behavior
        # and also captures Main breaker notes.
        possible_note = clean_token(remaining[-1])
        if possible_note.isdigit() and len(remaining) > 1:
            details["Notes"] = possible_note
            remaining = remaining[:-1]
        details["Lug Cable Size / Neut Sensor / Grnd Fault Cable"] = " ".join(
            remaining
        ).strip()

    return details


def breaker_family(catalog_number: str, abb_part_number: str) -> str:
    code = normalize_code(catalog_number or abb_part_number)
    xt_match = re.match(r"^(XT[1-7])", code)
    if xt_match:
        return xt_match.group(1)
    for prefix in ("TEY", "NEF", "1SDX", "1SQA"):
        if code.startswith(prefix):
            return prefix
    return "OTHER"


def _breaker_aggregate_key(row: dict) -> tuple:
    return tuple(
        row.get(column, "")
        for column in [
            "BOM Section",
            "Front View",
            "Sales Order",
            "Item",
            "Sales Order-Item",
            "Record Type",
            "Catalog Number",
            "Alternate SKU",
            "ABB Part Number",
            "Detected Code",
            "Mapping Found",
            "Source File",
            *BREAKER_DETAIL_FIELDS,
        ]
    )


def aggregate_breaker_occurrences(rows: Iterable[dict]) -> list[dict]:
    counter: Counter[tuple] = Counter()
    sample_by_key: dict[tuple, dict] = {}

    for row in rows:
        key = _breaker_aggregate_key(row)
        counter[key] += 1
        sample_by_key.setdefault(key, row)

    grouped: list[dict] = []
    for key, quantity in counter.items():
        row = dict(sample_by_key[key])
        row["Quantity"] = quantity
        grouped.append(row)
    return grouped


def _is_bom_stop_line(line: str, current_section: str) -> bool:
    upper = line.strip().upper()
    if not upper:
        return False
    if current_section == BOX_SECTION and "INTERIOR BOM" in upper:
        return True
    if upper in {"BOX BOM", "INTERIOR BOM"}:
        return True
    stop_prefixes = (
        "SYSTEM:",
        "QUANTITY:",
        "VOLTAGE:",
        "AMPS:",
        "PANEL TYPE:",
        "MATERIAL:",
        "KAIC:",
        "TECHNICAL SPECIFICATIONS",
        "MOUNTING:",
        "ENCLOSURE:",
        "SECTION:",
        "PLATE:",
        "PANEL INFORMATION",
        "INTERNAL DIMENSIONS",
        "EXTERNAL DIMENSIONS",
        "OPTIONS INCLUDED",
        "NOTES",
        "FEED DIR.",
        "TYPE:",
        "LUGS:",
        "AMPS/SENSOR:",
        "MAIN DISCONNECT DEVICE",
    )
    return upper.startswith(stop_prefixes)


def _clean_bom_description(description: str) -> str:
    text = description.strip()
    # PDF text extraction can append layout dimension labels to the last BOM
    # rows, for example: "1 XT1 Mtg Kit, Wide 2 in 2 in". Remove only
    # trailing inch measurements so real description text is preserved.
    text = re.sub(r"\s+\d+(?:\.\d+)?\s*in(?:\s+\d+(?:\.\d+)?\s*in)*\s*$", "", text, flags=re.IGNORECASE)
    return text.strip()


def _parse_bom_row(line: str) -> tuple[float, str, str] | None:
    cleaned = line.strip()
    if not cleaned:
        return None
    upper = cleaned.upper()
    if upper.startswith("QTY") or upper.startswith("CAT #"):
        return None

    match = re.match(
        r"^(-?\d+(?:\.\d+)?)\s+([A-Z0-9][A-Z0-9#./\-]{1,})\s+(.+?)\s*$",
        cleaned,
        re.IGNORECASE,
    )
    if not match:
        return None

    quantity_text, catalog_number, description = match.groups()
    description = _clean_bom_description(description)
    try:
        quantity = float(quantity_text)
    except ValueError:
        return None

    # Avoid layout dimensions accidentally parsed as catalog numbers.
    if len(catalog_number) < 3:
        return None
    if re.fullmatch(r"\d+(?:\.\d+)?", catalog_number):
        return None

    if quantity.is_integer():
        quantity = int(quantity)
    return quantity, normalize_code(catalog_number), description.strip()


def extract_bom_section_rows(
    text: str,
    section_name: str,
    front_view: str,
    source_file: str,
    page_number: int,
    sales_order: str,
    item: str,
    so_item: str,
    alt_to_abb: dict[str, str],
    abb_to_alt: dict[str, str],
) -> list[dict]:
    lines = [line.strip() for line in text.splitlines()]
    start_index = None
    target = section_name.upper()
    for index, line in enumerate(lines):
        if target in line.upper():
            start_index = index + 1
            break
    if start_index is None:
        return []

    rows: list[dict] = []
    for line in lines[start_index:]:
        if _is_bom_stop_line(line, section_name):
            break
        parsed = _parse_bom_row(line)
        if parsed is None:
            continue
        quantity, catalog_number, description = parsed
        match = map_catalog_number(catalog_number, alt_to_abb, abb_to_alt)
        rows.append(
            {
                "Sales Order": sales_order,
                "Item": item,
                "Sales Order-Item": so_item,
                "Front View": front_view,
                "BOM Section": section_name,
                "Record Type": "BOM Item",
                "Catalog Number": match.catalog_number,
                "Alternate SKU": match.alternate_sku,
                "ABB Part Number": match.abb_part_number,
                "Quantity": quantity,
                "Description": description,
                "Detected Code": match.detected_code,
                "Mapping Found": match.mapping_found,
                "Source File": source_file,
                "Page": page_number,
                "Original Line": line,
            }
        )
    return rows


def process_breaker_rows(
    uploaded_pdf: UploadedPDF,
    page_texts: list[str],
    alt_to_abb: dict[str, str],
    abb_to_alt: dict[str, str],
    sales_order: str,
    item: str,
    so_item: str,
) -> tuple[list[dict], list[dict], list[int]]:
    groups = group_breaker_pages(page_texts)
    actual_occurrences: list[dict] = []
    spacer_occurrences: list[dict] = []
    selected_pages: list[int] = []
    total_groups = len(groups)

    for group_index, page_group in enumerate(groups, start=1):
        front_view = panel_label(group_index, total_groups)
        selected_pages.extend(page_group)
        group_text = "\n".join(page_texts[index] for index in page_group)
        spacer_notes = identify_spacer_notes(group_text)

        for page_index in page_group:
            for line in page_texts[page_index].splitlines():
                matches = find_breaker_matches_in_line(line, alt_to_abb, abb_to_alt)
                if not matches:
                    continue

                row_note = note_at_end_of_row(line)
                is_spacer = bool(row_note and row_note in spacer_notes)

                for match in matches:
                    row_details = parse_breaker_row_details(line, match.detected_code)
                    row = {
                        "Sales Order": sales_order,
                        "Item": item,
                        "Sales Order-Item": so_item,
                        "Front View": front_view,
                        "BOM Section": BREAKER_SECTION,
                        "Record Type": "Spacer Breaker" if is_spacer else "Included Breaker",
                        "Catalog Number": match.catalog_number,
                        "Alternate SKU": match.alternate_sku,
                        "ABB Part Number": match.abb_part_number,
                        "Quantity": 1,
                        "Description": "",
                        "Detected Code": match.detected_code,
                        "Mapping Found": match.mapping_found,
                        "Source File": uploaded_pdf.name,
                        "Page": page_index + 1,
                        "Original Line": line.strip(),
                        **row_details,
                    }
                    if is_spacer:
                        spacer_occurrences.append(row)
                    else:
                        actual_occurrences.append(row)

    actual_grouped = aggregate_breaker_occurrences(actual_occurrences)
    spacer_grouped = aggregate_breaker_occurrences(spacer_occurrences)

    if spacer_grouped:
        details_by_key: dict[tuple, list[dict]] = {}
        for row in spacer_occurrences:
            key = _breaker_aggregate_key(row)
            details_by_key.setdefault(key, []).append(row)

        for row in spacer_grouped:
            key = _breaker_aggregate_key(row)
            details = details_by_key.get(key, [])
            row["Pages"] = ", ".join(
                sorted({str(detail["Page"]) for detail in details}, key=int)
            )
            row["Original Lines"] = " | ".join(
                detail["Original Line"] for detail in details
            )

    return actual_grouped, spacer_grouped, selected_pages


def process_bom_rows(
    uploaded_pdf: UploadedPDF,
    page_texts: list[str],
    alt_to_abb: dict[str, str],
    abb_to_alt: dict[str, str],
    sales_order: str,
    item: str,
    so_item: str,
    sections_to_extract: set[str],
) -> tuple[list[dict], list[int]]:
    bom_page_indexes = group_bom_pages(page_texts)
    bom_rows: list[dict] = []
    selected_pages: list[int] = []
    total_bom_pages = len(bom_page_indexes)

    for panel_index, page_index in enumerate(bom_page_indexes, start=1):
        text = page_texts[page_index]
        front_view = panel_label(panel_index, total_bom_pages)
        selected_pages.append(page_index)
        if BOX_SECTION in sections_to_extract:
            bom_rows.extend(
                extract_bom_section_rows(
                    text=text,
                    section_name=BOX_SECTION,
                    front_view=front_view,
                    source_file=uploaded_pdf.name,
                    page_number=page_index + 1,
                    sales_order=sales_order,
                    item=item,
                    so_item=so_item,
                    alt_to_abb=alt_to_abb,
                    abb_to_alt=abb_to_alt,
                )
            )
        if INTERIOR_SECTION in sections_to_extract:
            bom_rows.extend(
                extract_bom_section_rows(
                    text=text,
                    section_name=INTERIOR_SECTION,
                    front_view=front_view,
                    source_file=uploaded_pdf.name,
                    page_number=page_index + 1,
                    sales_order=sales_order,
                    item=item,
                    so_item=so_item,
                    alt_to_abb=alt_to_abb,
                    abb_to_alt=abb_to_alt,
                )
            )
    return bom_rows, selected_pages


def extract_pdf_text_pages(pdf_data: bytes) -> list[str]:
    """Extract page text using pypdf layout mode.

    pypdf is pure Python, so this avoids the native pypdfium/pdfplumber stack
    that was causing segmentation faults in Streamlit Community Cloud while
    still preserving table-like line layout for FV BOM pages.
    """
    reader = PdfReader(BytesIO(pdf_data))
    page_texts: list[str] = []
    for page in reader.pages:
        text = ""
        try:
            text = page.extract_text(extraction_mode="layout") or ""
        except Exception:
            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""
        page_texts.append(text)
    return page_texts or [""]


def process_pdf_bytes(
    uploaded_pdf: UploadedPDF,
    alt_to_abb: dict[str, str],
    abb_to_alt: dict[str, str],
    sections_to_extract: set[str],
) -> tuple[list[dict], list[dict], list[dict], dict]:
    sales_order, item, so_item = parse_so_item(uploaded_pdf.name)

    page_texts = extract_pdf_text_pages(uploaded_pdf.data)

    actual_breakers: list[dict] = []
    spacer_breakers: list[dict] = []
    bom_rows: list[dict] = []
    breaker_pages: list[int] = []
    bom_pages: list[int] = []

    if BREAKER_SECTION in sections_to_extract:
        actual_breakers, spacer_breakers, breaker_pages = process_breaker_rows(
            uploaded_pdf,
            page_texts,
            alt_to_abb,
            abb_to_alt,
            sales_order,
            item,
            so_item,
        )

    if BOX_SECTION in sections_to_extract or INTERIOR_SECTION in sections_to_extract:
        bom_rows, bom_pages = process_bom_rows(
            uploaded_pdf,
            page_texts,
            alt_to_abb,
            abb_to_alt,
            sales_order,
            item,
            so_item,
            sections_to_extract,
        )

    log = {
        "Source File": uploaded_pdf.name,
        "Sales Order": sales_order,
        "Item": item,
        "Sales Order-Item": so_item,
        "Breaker Table Pages": ", ".join(
            str(page + 1) for page in sorted(set(breaker_pages))
        ),
        "BOM Pages": ", ".join(str(page + 1) for page in sorted(set(bom_pages))),
        "Breaker Rows": sum(row["Quantity"] for row in actual_breakers),
        "Spacer Breakers Excluded": sum(row["Quantity"] for row in spacer_breakers),
        "Box BOM Rows": sum(1 for row in bom_rows if row["BOM Section"] == BOX_SECTION),
        "Interior BOM Rows": sum(1 for row in bom_rows if row["BOM Section"] == INTERIOR_SECTION),
        "Status": "Processed",
    }
    return actual_breakers, spacer_breakers, bom_rows, log


def sort_result(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    columns = [
        column
        for column in [
            "Sales Order",
            "Item",
            "Front View",
            "BOM Section",
            "Record Type",
            "Catalog Number",
            "ABB Part Number",
        ]
        if column in df.columns
    ]
    return df.sort_values(columns, kind="stable").reset_index(drop=True)


def _ensure_columns(dataframe: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    df = dataframe.copy()
    for column in columns:
        if column not in df.columns:
            df[column] = ""
    return df[list(columns)]


def build_dataframes(
    actual_breaker_rows: list[dict],
    spacer_breaker_rows: list[dict],
    bom_rows: list[dict],
    logs: list[dict],
    sections_to_extract: set[str],
) -> dict[str, pd.DataFrame]:
    base_columns = [
        "Sales Order",
        "Item",
        "Sales Order-Item",
        "Front View",
        "BOM Section",
        "Record Type",
        "Catalog Number",
        "Alternate SKU",
        "ABB Part Number",
        "Quantity",
        "Description",
        *BREAKER_DETAIL_FIELDS,
    ]
    internal_columns = base_columns + [
        "Detected Code",
        "Mapping Found",
        "Source File",
        "Page",
        "Original Line",
    ]
    spacer_columns = internal_columns + ["Pages", "Original Lines"]

    df_actual = sort_result(_ensure_columns(pd.DataFrame(actual_breaker_rows), internal_columns))
    df_spacers = sort_result(_ensure_columns(pd.DataFrame(spacer_breaker_rows), spacer_columns))
    df_bom = sort_result(_ensure_columns(pd.DataFrame(bom_rows), internal_columns))
    df_log = pd.DataFrame(logs)

    combined_internal = sort_result(
        pd.concat(
            [df_actual[internal_columns], df_spacers[internal_columns], df_bom[internal_columns]],
            ignore_index=True,
        )
    )
    df_results = combined_internal[base_columns].copy()

    if df_actual.empty:
        df_summary_family = pd.DataFrame(
            columns=[
                "Sales Order",
                "Item",
                "Sales Order-Item",
                "Front View",
                "Breaker Family",
                "Quantity",
            ]
        )
    else:
        df_family = df_actual.copy()
        df_family["Breaker Family"] = df_family.apply(
            lambda row: breaker_family(row["Catalog Number"], row["ABB Part Number"]),
            axis=1,
        )
        df_summary_family = (
            df_family.groupby(
                [
                    "Sales Order",
                    "Item",
                    "Sales Order-Item",
                    "Front View",
                    "Breaker Family",
                ],
                as_index=False,
            )["Quantity"]
            .sum()
            .sort_values(["Sales Order", "Item", "Front View", "Breaker Family"])
        )

    df_unmapped = combined_internal[~combined_internal["Mapping Found"].astype(bool)].copy()
    if df_unmapped.empty:
        df_unmapped = pd.DataFrame(
            columns=[
                "Detected Code",
                "BOM Section",
                "Record Type",
                "Sales Order-Item",
                "Front View",
                "Quantity",
                "Source File",
            ]
        )
    else:
        df_unmapped = df_unmapped[
            [
                "Detected Code",
                "BOM Section",
                "Record Type",
                "Sales Order-Item",
                "Front View",
                "Quantity",
                "Source File",
            ]
        ].copy()

    dataframes: dict[str, pd.DataFrame] = {
        "Results": df_results,
        "Unmapped Codes": df_unmapped,
        "Processing Log": df_log,
    }

    if BREAKER_SECTION in sections_to_extract:
        dataframes["Breaker BOM"] = df_actual[base_columns + ["Source File", "Page", "Original Line"]].copy()
        dataframes["Spacer Breakers"] = df_spacers[
            base_columns + ["Source File", "Pages", "Original Lines"]
        ].copy()
        dataframes["Summary by Breaker Family"] = df_summary_family
    if INTERIOR_SECTION in sections_to_extract:
        dataframes["Interior BOM"] = df_bom[df_bom["BOM Section"] == INTERIOR_SECTION][
            base_columns + ["Source File", "Page", "Original Line"]
        ].copy()
    if BOX_SECTION in sections_to_extract:
        dataframes["Box BOM"] = df_bom[df_bom["BOM Section"] == BOX_SECTION][
            base_columns + ["Source File", "Page", "Original Line"]
        ].copy()

    # Keep Results as the first sheet, then section-specific sheets, then support sheets.
    ordered: dict[str, pd.DataFrame] = {"Results": dataframes["Results"]}
    for name in ["Breaker BOM", "Spacer Breakers", "Summary by Breaker Family", "Interior BOM", "Box BOM"]:
        if name in dataframes:
            ordered[name] = dataframes[name]
    ordered["Unmapped Codes"] = dataframes["Unmapped Codes"]
    ordered["Processing Log"] = dataframes["Processing Log"]
    return ordered


def _format_standard_sheet(worksheet) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = min(max(max((len(value) for value in values), default=0) + 2, 10), 55)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _merge_same_values(
    worksheet,
    column: int,
    start_row: int,
    end_row: int,
    parent_columns: Sequence[int] = (),
) -> None:
    if end_row < start_row:
        return

    group_start = start_row
    for row in range(start_row + 1, end_row + 2):
        reached_end = row == end_row + 1
        same_value = False
        if not reached_end:
            same_value = worksheet.cell(row, column).value == worksheet.cell(group_start, column).value
            if same_value and parent_columns:
                same_value = all(
                    worksheet.cell(row, parent).value == worksheet.cell(group_start, parent).value
                    for parent in parent_columns
                )

        if reached_end or not same_value:
            if row - group_start > 1:
                worksheet.merge_cells(
                    start_row=group_start,
                    start_column=column,
                    end_row=row - 1,
                    end_column=column,
                )
            group_start = row


def _add_formatted_results_sheet(writer: pd.ExcelWriter, results: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    sheet_name = "Formatted Results"
    worksheet = writer.book.create_sheet(sheet_name, 0)
    headers = [
        "Sales Order",
        "Item",
        "Sales Order-Item",
        "Front View",
        "BOM Section",
        "Record Type",
        "Catalog Number",
        "ABB Part Number",
        "Quantity",
        "Description",
    ]

    for column, header in enumerate(headers, start=1):
        worksheet.cell(1, column, header)

    sorted_results = sort_result(results[headers].copy()) if not results.empty else results
    for row_number, values in enumerate(sorted_results.itertuples(index=False, name=None), start=2):
        for column, value in enumerate(values, start=1):
            worksheet.cell(row_number, column, value)

    header_fill = PatternFill("solid", fgColor="FF000F")
    header_font = Font(color="FFFFFF", bold=True)
    breaker_fill = PatternFill("solid", fgColor="E2F0D9")
    spacer_fill = PatternFill("solid", fgColor="FCE4D6")
    bom_fill = PatternFill("solid", fgColor="EEF2F7")
    thin_gray = Side(style="thin", color="D9E1F2")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = worksheet.max_row
    for row in range(2, last_row + 1):
        record_type = worksheet.cell(row, 6).value
        section = worksheet.cell(row, 5).value
        if record_type == "Spacer Breaker":
            row_fill = spacer_fill
        elif section == BREAKER_SECTION:
            row_fill = breaker_fill
        else:
            row_fill = bom_fill
        for column in range(1, len(headers) + 1):
            cell = worksheet.cell(row, column)
            cell.fill = row_fill
            cell.border = Border(bottom=thin_gray)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    if last_row >= 2:
        _merge_same_values(worksheet, 6, 2, last_row, parent_columns=(1, 2, 3, 4, 5))
        _merge_same_values(worksheet, 5, 2, last_row, parent_columns=(1, 2, 3, 4))
        _merge_same_values(worksheet, 4, 2, last_row, parent_columns=(1, 2, 3))
        _merge_same_values(worksheet, 3, 2, last_row, parent_columns=(1, 2))
        _merge_same_values(worksheet, 2, 2, last_row, parent_columns=(1,))
        _merge_same_values(worksheet, 1, 2, last_row)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row[:6]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [16, 10, 22, 22, 18, 18, 32, 24, 10, 45]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width

    worksheet.freeze_panes = "G2"
    worksheet.sheet_view.showGridLines = False


def dataframes_to_excel(dataframes: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dataframe in dataframes.items():
            dataframe.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            _format_standard_sheet(writer.book[sheet_name[:31]])
        _add_formatted_results_sheet(writer, dataframes["Results"])
    output.seek(0)
    return output.getvalue()


def process_uploaded_pdfs(
    pdfs: Sequence[UploadedPDF],
    mapping_data: bytes,
    manual_mappings: Mapping[str, str] | None = None,
    progress_callback: ProgressCallback | None = None,
    sections_to_extract: Iterable[str] | None = None,
    generate_excel: bool = True,
) -> tuple[dict[str, pd.DataFrame], bytes, dict[str, int]]:
    selected_sections = set(sections_to_extract or ALL_SECTIONS)
    selected_sections = selected_sections.intersection(set(ALL_SECTIONS))
    if not selected_sections:
        selected_sections = {BREAKER_SECTION}

    alt_to_abb, abb_to_alt = load_mappings_from_bytes(mapping_data, manual_mappings=manual_mappings)

    all_actual_breakers: list[dict] = []
    all_spacer_breakers: list[dict] = []
    all_bom_rows: list[dict] = []
    logs: list[dict] = []
    total = len(pdfs)

    for index, pdf in enumerate(pdfs, start=1):
        try:
            actual, spacers, bom_rows, log = process_pdf_bytes(
                pdf,
                alt_to_abb,
                abb_to_alt,
                selected_sections,
            )
            all_actual_breakers.extend(actual)
            all_spacer_breakers.extend(spacers)
            all_bom_rows.extend(bom_rows)
            logs.append(log)
        except Exception as exc:
            sales_order, item, so_item = parse_so_item(pdf.name)
            logs.append(
                {
                    "Source File": pdf.name,
                    "Sales Order": sales_order,
                    "Item": item,
                    "Sales Order-Item": so_item,
                    "Breaker Table Pages": "",
                    "BOM Pages": "",
                    "Breaker Rows": 0,
                    "Spacer Breakers Excluded": 0,
                    "Box BOM Rows": 0,
                    "Interior BOM Rows": 0,
                    "Status": f"ERROR: {exc}",
                }
            )
        finally:
            if progress_callback:
                progress_callback(index, total, pdf.name)

    dataframes = build_dataframes(
        all_actual_breakers,
        all_spacer_breakers,
        all_bom_rows,
        logs,
        selected_sections,
    )
    excel_bytes = dataframes_to_excel(dataframes) if generate_excel else b""

    unique_so_items = {
        row["Sales Order-Item"]
        for row in all_actual_breakers + all_spacer_breakers + all_bom_rows
    }
    unmapped_df = dataframes["Unmapped Codes"]
    metrics = {
        "pdfs": len(pdfs),
        "so_items": len(unique_so_items),
        "breaker_rows": sum(row["Quantity"] for row in all_actual_breakers),
        "spacer_rows": sum(row["Quantity"] for row in all_spacer_breakers),
        "box_bom_rows": sum(1 for row in all_bom_rows if row["BOM Section"] == BOX_SECTION),
        "interior_bom_rows": sum(1 for row in all_bom_rows if row["BOM Section"] == INTERIOR_SECTION),
        "unmapped": unmapped_df["Detected Code"].nunique() if not unmapped_df.empty else 0,
        "errors": sum(1 for row in logs if str(row["Status"]).startswith("ERROR")),
    }
    return dataframes, excel_bytes, metrics
